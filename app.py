#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║                 CALCULATEUR D'INTÉRÊTS DE RETARD                             ║
║                       Eurovia / VINCI Construction                           ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  Application Streamlit professionnelle pour le calcul des intérêts          ║
║  moratoires selon :                                                          ║
║  • L441-10 C.Com (Clients Privés) : BCE + 10 pts, semestriel                ║
║  • R2192-31 CCP (Clients Publics) : BCE + 8 pts, annuel (1er janvier)       ║
║  • Taux Manuel/Contractuel                                                   ║
╚══════════════════════════════════════════════════════════════════════════════╝

Version: 2.0.0
Date: Janvier 2026
"""

import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional, List, Dict, Any, Tuple
import io
import csv
import ssl
import urllib.request
import base64
from dataclasses import dataclass
from enum import Enum
import certifi
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION DE LA PAGE
# ═══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Calculateur Intérêts de Retard | Eurovia",
    page_icon="💶",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ═══════════════════════════════════════════════════════════════════════════════
# STYLES CSS - DESIGN INSPIRÉ APPLE / VINCI
# ═══════════════════════════════════════════════════════════════════════════════

CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=SF+Pro+Display:wght@300;400;500;600;700&family=Inter:wght@300;400;500;600;700&display=swap');

:root {
    --vinci-blue: #0B3DAA;
    --vinci-red: #7B1626;
    --eurovia-blue: #003366;
    --background: #F5F7FA;
    --surface: #FFFFFF;
    --text-primary: #1D1D1F;
    --text-secondary: #6B7280;
    --border: #E5E7EB;
    --success: #10B981;
    --warning: #F59E0B;
    --error: #EF4444;
    --radius-sm: 8px;
    --radius-md: 12px;
    --radius-lg: 16px;
    --shadow-sm: 0 1px 2px rgba(0,0,0,0.05);
    --shadow-md: 0 4px 6px -1px rgba(0,0,0,0.1), 0 2px 4px -1px rgba(0,0,0,0.06);
    --shadow-lg: 0 10px 15px -3px rgba(0,0,0,0.1), 0 4px 6px -2px rgba(0,0,0,0.05);
}

/* Reset et base */
.stApp {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    background: linear-gradient(135deg, #F8FAFC 0%, #EEF2F7 100%);
}

/* Header principal */
.main-header {
    background: linear-gradient(135deg, var(--vinci-blue) 0%, #1E4FC2 100%);
    padding: 2rem;
    border-radius: var(--radius-lg);
    margin-bottom: 2rem;
    box-shadow: var(--shadow-lg);
    position: relative;
    overflow: hidden;
}

.main-header::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -20%;
    width: 60%;
    height: 200%;
    background: linear-gradient(135deg, rgba(255,255,255,0.1) 0%, transparent 60%);
    transform: rotate(-15deg);
}

.main-header h1 {
    color: white;
    font-weight: 700;
    font-size: 2rem;
    margin: 0;
    position: relative;
    z-index: 1;
}

.main-header p {
    color: rgba(255,255,255,0.85);
    font-size: 1rem;
    margin-top: 0.5rem;
    position: relative;
    z-index: 1;
}

/* Cards */
.card {
    background: var(--surface);
    border-radius: var(--radius-md);
    padding: 1.5rem;
    box-shadow: var(--shadow-md);
    border: 1px solid var(--border);
    margin-bottom: 1rem;
    transition: transform 0.2s ease, box-shadow 0.2s ease;
}

.card:hover {
    transform: translateY(-2px);
    box-shadow: var(--shadow-lg);
}

.card-title {
    font-size: 1.1rem;
    font-weight: 600;
    color: var(--vinci-blue);
    margin-bottom: 1rem;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}

/* Result cards */
.result-card {
    background: linear-gradient(135deg, var(--vinci-blue) 0%, #1E4FC2 100%);
    color: white;
    border-radius: var(--radius-md);
    padding: 1.5rem;
    text-align: center;
    box-shadow: var(--shadow-lg);
}

.result-card.success {
    background: linear-gradient(135deg, var(--success) 0%, #059669 100%);
}

.result-card.warning {
    background: linear-gradient(135deg, var(--warning) 0%, #D97706 100%);
}

.result-value {
    font-size: 2rem;
    font-weight: 700;
    margin: 0.5rem 0;
}

.result-label {
    font-size: 0.875rem;
    opacity: 0.9;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}

/* Input styling */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stDateInput > div > div > input,
.stSelectbox > div > div {
    border-radius: var(--radius-sm) !important;
    border: 2px solid var(--border) !important;
    padding: 0.75rem !important;
    font-size: 1rem !important;
    transition: border-color 0.2s ease, box-shadow 0.2s ease !important;
}

.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus,
.stDateInput > div > div > input:focus {
    border-color: var(--vinci-blue) !important;
    box-shadow: 0 0 0 3px rgba(11, 61, 170, 0.1) !important;
}

/* Buttons */
.stButton > button {
    background: linear-gradient(135deg, var(--vinci-blue) 0%, #1E4FC2 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: var(--radius-sm) !important;
    padding: 0.75rem 2rem !important;
    font-weight: 600 !important;
    font-size: 1rem !important;
    transition: transform 0.2s ease, box-shadow 0.2s ease !important;
    box-shadow: var(--shadow-md) !important;
}

.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: var(--shadow-lg) !important;
}

.stButton > button:active {
    transform: translateY(0) !important;
}

/* Secondary button */
.secondary-btn > button {
    background: var(--surface) !important;
    color: var(--vinci-blue) !important;
    border: 2px solid var(--vinci-blue) !important;
}

/* Radio buttons */
.stRadio > div {
    gap: 1rem;
}

.stRadio > div > label {
    background: var(--surface);
    padding: 1rem 1.5rem;
    border-radius: var(--radius-sm);
    border: 2px solid var(--border);
    cursor: pointer;
    transition: all 0.2s ease;
}

.stRadio > div > label:hover {
    border-color: var(--vinci-blue);
}

.stRadio > div > label[data-checked="true"] {
    border-color: var(--vinci-blue);
    background: rgba(11, 61, 170, 0.05);
}

/* Tables */
.dataframe {
    border-radius: var(--radius-md) !important;
    overflow: hidden;
    box-shadow: var(--shadow-sm);
}

.dataframe thead th {
    background: var(--vinci-blue) !important;
    color: white !important;
    font-weight: 600 !important;
    padding: 1rem !important;
    text-transform: uppercase;
    font-size: 0.75rem;
    letter-spacing: 0.05em;
}

.dataframe tbody td {
    padding: 0.875rem 1rem !important;
    border-bottom: 1px solid var(--border) !important;
}

.dataframe tbody tr:hover {
    background: rgba(11, 61, 170, 0.02) !important;
}

/* Sidebar */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #FFFFFF 0%, #F8FAFC 100%);
    border-right: 1px solid var(--border);
}

[data-testid="stSidebar"] .stMarkdown h3 {
    color: var(--vinci-blue);
    font-weight: 600;
}

/* Metrics */
[data-testid="stMetricValue"] {
    font-size: 1.75rem !important;
    font-weight: 700 !important;
    color: var(--vinci-blue) !important;
}

[data-testid="stMetricLabel"] {
    font-size: 0.875rem !important;
    color: var(--text-secondary) !important;
    text-transform: uppercase;
    letter-spacing: 0.03em;
}

/* Alerts */
.stAlert {
    border-radius: var(--radius-sm) !important;
    border: none !important;
}

/* Expander */
.streamlit-expanderHeader {
    background: var(--surface) !important;
    border-radius: var(--radius-sm) !important;
    border: 1px solid var(--border) !important;
    font-weight: 600 !important;
    color: var(--vinci-blue) !important;
}

/* Divider */
hr {
    border: none;
    height: 1px;
    background: linear-gradient(90deg, transparent, var(--border), transparent);
    margin: 2rem 0;
}

/* Logo container */
.logo-container {
    display: flex;
    align-items: center;
    gap: 1rem;
    padding: 1rem;
    background: white;
    border-radius: var(--radius-md);
    margin-bottom: 1.5rem;
}

/* Info box */
.info-box {
    background: linear-gradient(135deg, rgba(11, 61, 170, 0.05) 0%, rgba(11, 61, 170, 0.02) 100%);
    border-left: 4px solid var(--vinci-blue);
    padding: 1rem 1.5rem;
    border-radius: 0 var(--radius-sm) var(--radius-sm) 0;
    margin: 1rem 0;
}

.info-box p {
    margin: 0;
    color: var(--text-primary);
    font-size: 0.9rem;
}

/* Status badge */
.status-badge {
    display: inline-flex;
    align-items: center;
    gap: 0.375rem;
    padding: 0.375rem 0.75rem;
    border-radius: 9999px;
    font-size: 0.75rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.03em;
}

.status-badge.success {
    background: rgba(16, 185, 129, 0.1);
    color: var(--success);
}

.status-badge.pending {
    background: rgba(245, 158, 11, 0.1);
    color: var(--warning);
}

/* Animation */
@keyframes fadeIn {
    from { opacity: 0; transform: translateY(10px); }
    to { opacity: 1; transform: translateY(0); }
}

.fade-in {
    animation: fadeIn 0.5s ease forwards;
}

/* Hide Streamlit branding */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# CLASSES ET TYPES
# ═══════════════════════════════════════════════════════════════════════════════

class ClientType(Enum):
    """Type de client pour le calcul des intérêts."""
    PRIVE = "prive"
    PUBLIC = "public"


class TauxMode(Enum):
    """Mode de calcul du taux d'intérêt."""
    L441_10 = "l441_10"  # Privé: BCE + 10 pts, semestriel
    R2192_31 = "r2192_31"  # Public: BCE + 8 pts, annuel
    MANUEL = "manuel"  # Taux contractuel fixe


@dataclass
class SegmentInteret:
    """Représente un segment de calcul d'intérêts."""
    debut: date
    fin: date
    jours: int
    taux_bce: float
    majoration: float
    taux_applique: float
    interets: Decimal
    periode_label: str


@dataclass
class ResultatCalcul:
    """Résultat complet d'un calcul d'intérêts."""
    montant_principal: Decimal
    date_echeance: date
    date_paiement: date
    jours_retard: int
    segments: List[SegmentInteret]
    interets_totaux: Decimal
    indemnite_forfaitaire: Decimal
    total_du: Decimal
    client_type: ClientType
    taux_mode: TauxMode
    reference_legale: str


# ═══════════════════════════════════════════════════════════════════════════════
# FONCTIONS UTILITAIRES
# ═══════════════════════════════════════════════════════════════════════════════

def round_decimal(value: Decimal, decimals: int = 2) -> Decimal:
    """Arrondi bancaire standard."""
    return value.quantize(Decimal(f"0.{'0' * decimals}"), rounding=ROUND_HALF_UP)


def format_currency(value: Decimal) -> str:
    """Formate un montant en euros (format français)."""
    rounded = round_decimal(value)
    formatted = f"{rounded:,.2f}".replace(",", " ").replace(".", ",")
    return f"{formatted} €"


def format_percentage(value: float) -> str:
    """Formate un pourcentage."""
    return f"{value:.2f} %".replace(".", ",")


def diff_days(start: date, end: date) -> int:
    """Calcule le nombre de jours entre deux dates."""
    return max(0, (end - start).days)


def get_semester_start(d: date) -> date:
    """Retourne le 1er janvier ou 1er juillet de l'année selon le semestre."""
    return date(d.year, 1, 1) if d.month <= 6 else date(d.year, 7, 1)


def get_next_semester_start(anchor: date) -> date:
    """Retourne la prochaine borne semestrielle."""
    if anchor.month == 1:
        return date(anchor.year, 7, 1)
    return date(anchor.year + 1, 1, 1)


def get_year_start(d: date) -> date:
    """Retourne le 1er janvier de l'année."""
    return date(d.year, 1, 1)


def get_next_year_start(d: date) -> date:
    """Retourne le 1er janvier de l'année suivante."""
    return date(d.year + 1, 1, 1)


def format_semester_label(anchor: date) -> str:
    """Retourne 'S1 YYYY' ou 'S2 YYYY'."""
    sem = "S1" if anchor.month == 1 else "S2"
    return f"{sem} {anchor.year}"


# ═══════════════════════════════════════════════════════════════════════════════
# TÉLÉCHARGEMENT DES TAUX BCE
# ═══════════════════════════════════════════════════════════════════════════════

def http_get_text(url: str, timeout: int = 20) -> str:
    """Requête HTTPS robuste avec gestion SSL."""
    # Bypass SSL verification for internal/corporate networks
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    
    handlers = []
    proxies = urllib.request.getproxies()
    if proxies:
        handlers.append(urllib.request.ProxyHandler(proxies))
    handlers.append(urllib.request.HTTPSHandler(context=ctx))
    
    opener = urllib.request.build_opener(*handlers)
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Eurovia-InterestCalc/2.0 (Python Streamlit)"}
    )
    
    with opener.open(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8", errors="replace")


def parse_date_flexible(s: str) -> date:
    """Parse une date avec plusieurs formats possibles."""
    s = s.strip()
    formats = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Date invalide: {s}")


@st.cache_data(ttl=3600, show_spinner=False)
def load_bce_rates_from_ecb() -> List[Dict[str, Any]]:
    """Charge les taux BCE depuis l'API de la BCE."""
    url = (
        "https://data-api.ecb.europa.eu/service/data/"
        "FM/D.U2.EUR.4F.KR.MRR_FR.LEV?detail=dataonly&format=csvdata"
    )
    
    text = http_get_text(url)
    
    # Détection du séparateur
    sample = "\n".join(text.splitlines()[:5])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ','
    
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    dates_vals = []
    
    for row in reader:
        d = row.get('TIME_PERIOD') or row.get('TIME') or row.get('Period') or row.get('DATE')
        v = row.get('OBS_VALUE') or row.get('OBS_VALUE (OBS_VALUE)') or row.get('VALUE') or row.get('OBS')
        
        if not d or not v:
            continue
        
        v_str = str(v).strip()
        if v_str in ('.', '', 'NaN', 'nan'):
            continue
        
        try:
            val = float(v_str.replace(',', '.'))
            dt = parse_date_flexible(d)
            dates_vals.append((dt, val))
        except (ValueError, TypeError):
            continue
    
    if not dates_vals:
        raise ValueError("Aucune donnée MRO trouvée dans la réponse BCE.")
    
    dates_vals.sort(key=lambda x: x[0])
    
    # Créer le schedule avec une entrée à chaque changement de taux
    schedule = []
    prev_val = None
    for dt, val in dates_vals:
        if prev_val is None or val != prev_val:
            schedule.append({'start': dt, 'rate': val})
        prev_val = val
    
    return schedule


@st.cache_data(ttl=3600, show_spinner=False)
def load_bce_rates_from_fred() -> List[Dict[str, Any]]:
    """Charge les taux BCE depuis FRED (fallback)."""
    url = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=ECBMRRFR"
    text = http_get_text(url)
    
    reader = csv.DictReader(io.StringIO(text))
    daily = []
    
    for row in reader:
        d = row.get('DATE')
        v = row.get('ECBMRRFR')
        
        if not d or not v:
            continue
        
        v_str = str(v).strip()
        if v_str in ('.', '', 'NaN', 'nan'):
            continue
        
        try:
            dt = parse_date_flexible(d)
            val = float(v_str)
            daily.append((dt, val))
        except (ValueError, TypeError):
            continue
    
    if not daily:
        raise ValueError("Aucune donnée MRO trouvée depuis FRED.")
    
    daily.sort(key=lambda x: x[0])
    
    schedule = []
    prev_val = None
    for dt, val in daily:
        if prev_val is None or val != prev_val:
            schedule.append({'start': dt, 'rate': val})
        prev_val = val
    
    return schedule


def get_bce_rates() -> List[Dict[str, Any]]:
    """Récupère les taux BCE avec fallback automatique."""
    try:
        return load_bce_rates_from_ecb()
    except Exception as e_ecb:
        try:
            return load_bce_rates_from_fred()
        except Exception as e_fred:
            st.error(f"Impossible de charger les taux BCE.\nBCE: {e_ecb}\nFRED: {e_fred}")
            return []


def get_rate_at_date(d: date, schedule: List[Dict[str, Any]]) -> float:
    """Retourne le taux applicable à une date donnée."""
    if not schedule:
        raise ValueError("Schedule vide")
    
    applicable_rate = schedule[0]['rate']
    for entry in schedule:
        if entry['start'] <= d:
            applicable_rate = entry['rate']
        else:
            break
    
    return applicable_rate


# ═══════════════════════════════════════════════════════════════════════════════
# CALCUL DES INTÉRÊTS
# ═══════════════════════════════════════════════════════════════════════════════

def calculer_interets_prive_l441_10(
    montant: Decimal,
    date_echeance: date,
    date_paiement: date,
    bce_rates: List[Dict[str, Any]]
) -> ResultatCalcul:
    """
    Calcul des intérêts pour clients PRIVÉS selon L441-10 C.Com.
    Taux = BCE + 10 points, actualisé par semestre (1er janvier / 1er juillet).
    """
    if date_paiement <= date_echeance:
        return ResultatCalcul(
            montant_principal=montant,
            date_echeance=date_echeance,
            date_paiement=date_paiement,
            jours_retard=0,
            segments=[],
            interets_totaux=Decimal('0'),
            indemnite_forfaitaire=Decimal('40.00'),
            total_du=Decimal('40.00'),
            client_type=ClientType.PRIVE,
            taux_mode=TauxMode.L441_10,
            reference_legale="Article L.441-10 du Code de commerce"
        )
    
    segments = []
    cursor = date_echeance
    total_interets = Decimal('0')
    majoration = 10.0  # +10 points pour le privé
    
    while cursor < date_paiement:
        # Déterminer la borne semestrielle
        anchor = get_semester_start(cursor)
        next_sem = get_next_semester_start(anchor)
        
        # Fin du segment = min(fin semestre, date paiement)
        seg_end = min(next_sem, date_paiement)
        
        # Taux BCE à la borne semestrielle
        taux_bce = get_rate_at_date(anchor, bce_rates)
        taux_applique = taux_bce + majoration
        
        # Calcul des jours et intérêts
        jours = diff_days(cursor, seg_end)
        if jours > 0:
            interet_segment = montant * Decimal(str(taux_applique)) / Decimal('100') * Decimal(str(jours)) / Decimal('365')
            interet_segment = round_decimal(interet_segment)
            
            segments.append(SegmentInteret(
                debut=cursor,
                fin=seg_end - timedelta(days=1),
                jours=jours,
                taux_bce=taux_bce,
                majoration=majoration,
                taux_applique=taux_applique,
                interets=interet_segment,
                periode_label=format_semester_label(anchor)
            ))
            
            total_interets += interet_segment
        
        cursor = seg_end
    
    total_interets = round_decimal(total_interets)
    indemnite = Decimal('40.00')
    total = round_decimal(total_interets + indemnite)
    
    return ResultatCalcul(
        montant_principal=montant,
        date_echeance=date_echeance,
        date_paiement=date_paiement,
        jours_retard=diff_days(date_echeance, date_paiement),
        segments=segments,
        interets_totaux=total_interets,
        indemnite_forfaitaire=indemnite,
        total_du=total,
        client_type=ClientType.PRIVE,
        taux_mode=TauxMode.L441_10,
        reference_legale="Article L.441-10 du Code de commerce"
    )


def calculer_interets_public_r2192_31(
    montant: Decimal,
    date_echeance: date,
    date_paiement: date,
    bce_rates: List[Dict[str, Any]]
) -> ResultatCalcul:
    """
    Calcul des intérêts pour clients PUBLICS selon R2192-31 CCP.
    Taux = BCE + 8 points, actualisé annuellement (1er janvier).
    """
    # Les intérêts commencent le lendemain de l'échéance
    debut_interets = date_echeance + timedelta(days=1)
    
    if date_paiement <= date_echeance:
        return ResultatCalcul(
            montant_principal=montant,
            date_echeance=date_echeance,
            date_paiement=date_paiement,
            jours_retard=0,
            segments=[],
            interets_totaux=Decimal('0'),
            indemnite_forfaitaire=Decimal('40.00'),
            total_du=Decimal('40.00'),
            client_type=ClientType.PUBLIC,
            taux_mode=TauxMode.R2192_31,
            reference_legale="Article R.2192-31 du Code de la commande publique"
        )
    
    segments = []
    cursor = debut_interets
    total_interets = Decimal('0')
    majoration = 8.0  # +8 points pour le public
    
    while cursor < date_paiement:
        # Borne annuelle (1er janvier)
        anchor = get_year_start(cursor)
        next_year = get_next_year_start(cursor)
        
        # Fin du segment = min(1er janvier année suivante, date paiement)
        seg_end = min(next_year, date_paiement)
        
        # Taux BCE au 1er janvier de l'année
        taux_bce = get_rate_at_date(anchor, bce_rates)
        taux_applique = taux_bce + majoration
        
        # Calcul des jours et intérêts
        jours = diff_days(cursor, seg_end)
        if jours > 0:
            interet_segment = montant * Decimal(str(taux_applique)) / Decimal('100') * Decimal(str(jours)) / Decimal('365')
            interet_segment = round_decimal(interet_segment)
            
            segments.append(SegmentInteret(
                debut=cursor,
                fin=seg_end - timedelta(days=1) if seg_end != date_paiement else date_paiement,
                jours=jours,
                taux_bce=taux_bce,
                majoration=majoration,
                taux_applique=taux_applique,
                interets=interet_segment,
                periode_label=f"Année {anchor.year}"
            ))
            
            total_interets += interet_segment
        
        cursor = seg_end
    
    total_interets = round_decimal(total_interets)
    indemnite = Decimal('40.00')
    total = round_decimal(total_interets + indemnite)
    
    return ResultatCalcul(
        montant_principal=montant,
        date_echeance=date_echeance,
        date_paiement=date_paiement,
        jours_retard=diff_days(date_echeance, date_paiement),
        segments=segments,
        interets_totaux=total_interets,
        indemnite_forfaitaire=indemnite,
        total_du=total,
        client_type=ClientType.PUBLIC,
        taux_mode=TauxMode.R2192_31,
        reference_legale="Article R.2192-31 du Code de la commande publique"
    )


def calculer_interets_manuel(
    montant: Decimal,
    date_echeance: date,
    date_paiement: date,
    taux_contractuel: float
) -> ResultatCalcul:
    """
    Calcul des intérêts avec un taux contractuel fixe (manuel).
    """
    if date_paiement <= date_echeance:
        return ResultatCalcul(
            montant_principal=montant,
            date_echeance=date_echeance,
            date_paiement=date_paiement,
            jours_retard=0,
            segments=[],
            interets_totaux=Decimal('0'),
            indemnite_forfaitaire=Decimal('40.00'),
            total_du=Decimal('40.00'),
            client_type=ClientType.PRIVE,
            taux_mode=TauxMode.MANUEL,
            reference_legale="Clause contractuelle"
        )
    
    jours = diff_days(date_echeance, date_paiement)
    
    interets = montant * Decimal(str(taux_contractuel)) / Decimal('100') * Decimal(str(jours)) / Decimal('365')
    interets = round_decimal(interets)
    
    segment = SegmentInteret(
        debut=date_echeance,
        fin=date_paiement - timedelta(days=1),
        jours=jours,
        taux_bce=0.0,
        majoration=0.0,
        taux_applique=taux_contractuel,
        interets=interets,
        periode_label="Taux contractuel"
    )
    
    indemnite = Decimal('40.00')
    total = round_decimal(interets + indemnite)
    
    return ResultatCalcul(
        montant_principal=montant,
        date_echeance=date_echeance,
        date_paiement=date_paiement,
        jours_retard=jours,
        segments=[segment],
        interets_totaux=interets,
        indemnite_forfaitaire=indemnite,
        total_du=total,
        client_type=ClientType.PRIVE,
        taux_mode=TauxMode.MANUEL,
        reference_legale="Clause contractuelle"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# GÉNÉRATION EXPORTS
# ═══════════════════════════════════════════════════════════════════════════════

def generate_html_report(resultat: ResultatCalcul, libelle: str) -> str:
    """Génère un rapport HTML détaillé."""
    
    # Texte de référence légale
    if resultat.taux_mode == TauxMode.L441_10:
        ref_text = """
        <strong>Article L.441-10 du Code de commerce :</strong><br>
        Les pénalités de retard sont exigibles sans qu'un rappel soit nécessaire. 
        Le taux des pénalités est égal au taux d'intérêt appliqué par la Banque centrale européenne 
        à son opération de refinancement la plus récente majoré de <strong>10 points de pourcentage</strong>. 
        Le taux applicable pendant le premier semestre de l'année concernée est le taux en vigueur 
        au 1er janvier ; pour le second semestre, celui en vigueur au 1er juillet.
        """
    elif resultat.taux_mode == TauxMode.R2192_31:
        ref_text = """
        <strong>Article R.2192-31 du Code de la commande publique :</strong><br>
        Le taux des intérêts moratoires est égal au taux d'intérêt de la principale facilité 
        de refinancement appliquée par la Banque centrale européenne à son opération de 
        refinancement principal la plus récente effectuée avant le premier jour de calendrier 
        du semestre de l'année civile au cours duquel les intérêts moratoires ont commencé à courir, 
        majoré de <strong>8 points de pourcentage</strong>.
        """
    else:
        ref_text = """
        <strong>Clause contractuelle :</strong><br>
        Le taux d'intérêt de retard appliqué est celui défini contractuellement entre les parties.
        """
    
    # Lignes de segments
    rows_html = ""
    for seg in resultat.segments:
        rows_html += f"""
        <tr>
            <td>{seg.debut.strftime('%d/%m/%Y')} → {seg.fin.strftime('%d/%m/%Y')}</td>
            <td class="center">{seg.jours}</td>
            <td class="center">{seg.taux_bce:.2f} %</td>
            <td class="center">+{seg.majoration:.0f} pts</td>
            <td class="center">{seg.taux_applique:.2f} %</td>
            <td class="right">{format_currency(seg.interets)}</td>
        </tr>
        """
    
    html = f"""
    <!DOCTYPE html>
    <html lang="fr">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Calcul Intérêts de Retard - {libelle}</title>
        <style>
            * {{
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }}
            body {{
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
                line-height: 1.6;
                color: #1D1D1F;
                background: #F5F7FA;
                padding: 2rem;
            }}
            .container {{
                max-width: 900px;
                margin: 0 auto;
                background: white;
                border-radius: 16px;
                box-shadow: 0 4px 24px rgba(0,0,0,0.1);
                overflow: hidden;
            }}
            .header {{
                background: linear-gradient(135deg, #0B3DAA 0%, #1E4FC2 100%);
                color: white;
                padding: 2rem;
            }}
            .header h1 {{
                font-size: 1.75rem;
                font-weight: 700;
                margin-bottom: 0.5rem;
            }}
            .header p {{
                opacity: 0.9;
            }}
            .content {{
                padding: 2rem;
            }}
            .section {{
                margin-bottom: 2rem;
            }}
            .section-title {{
                color: #0B3DAA;
                font-size: 1.1rem;
                font-weight: 600;
                margin-bottom: 1rem;
                padding-bottom: 0.5rem;
                border-bottom: 2px solid #E5E7EB;
            }}
            .info-grid {{
                display: grid;
                grid-template-columns: repeat(2, 1fr);
                gap: 1rem;
            }}
            .info-item {{
                background: #F8FAFC;
                padding: 1rem;
                border-radius: 8px;
            }}
            .info-item .label {{
                font-size: 0.75rem;
                color: #6B7280;
                text-transform: uppercase;
                letter-spacing: 0.05em;
            }}
            .info-item .value {{
                font-size: 1.25rem;
                font-weight: 600;
                color: #1D1D1F;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
            }}
            th {{
                background: #0B3DAA;
                color: white;
                padding: 0.875rem 1rem;
                text-align: left;
                font-size: 0.75rem;
                text-transform: uppercase;
                letter-spacing: 0.05em;
            }}
            td {{
                padding: 0.875rem 1rem;
                border-bottom: 1px solid #E5E7EB;
            }}
            .center {{ text-align: center; }}
            .right {{ text-align: right; }}
            .ref-box {{
                background: linear-gradient(135deg, rgba(11, 61, 170, 0.05) 0%, rgba(11, 61, 170, 0.02) 100%);
                border-left: 4px solid #0B3DAA;
                padding: 1rem 1.5rem;
                border-radius: 0 8px 8px 0;
                font-size: 0.9rem;
            }}
            .total-box {{
                background: linear-gradient(135deg, #0B3DAA 0%, #1E4FC2 100%);
                color: white;
                padding: 1.5rem;
                border-radius: 12px;
                text-align: center;
            }}
            .total-box .amount {{
                font-size: 2.5rem;
                font-weight: 700;
            }}
            .total-box .label {{
                opacity: 0.9;
                text-transform: uppercase;
                letter-spacing: 0.05em;
                font-size: 0.875rem;
            }}
            .footer {{
                background: #F8FAFC;
                padding: 1.5rem 2rem;
                text-align: center;
                font-size: 0.875rem;
                color: #6B7280;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Calcul des Intérêts de Retard</h1>
                <p>{libelle}</p>
            </div>
            
            <div class="content">
                <div class="section">
                    <h2 class="section-title">📋 Informations de la facture</h2>
                    <div class="info-grid">
                        <div class="info-item">
                            <div class="label">Montant principal</div>
                            <div class="value">{format_currency(resultat.montant_principal)}</div>
                        </div>
                        <div class="info-item">
                            <div class="label">Type de client</div>
                            <div class="value">{'Privé' if resultat.client_type == ClientType.PRIVE else 'Public'}</div>
                        </div>
                        <div class="info-item">
                            <div class="label">Date d'échéance</div>
                            <div class="value">{resultat.date_echeance.strftime('%d/%m/%Y')}</div>
                        </div>
                        <div class="info-item">
                            <div class="label">Date de paiement</div>
                            <div class="value">{resultat.date_paiement.strftime('%d/%m/%Y')}</div>
                        </div>
                        <div class="info-item">
                            <div class="label">Jours de retard</div>
                            <div class="value">{resultat.jours_retard} jours</div>
                        </div>
                        <div class="info-item">
                            <div class="label">Base légale</div>
                            <div class="value">{resultat.reference_legale}</div>
                        </div>
                    </div>
                </div>
                
                <div class="section">
                    <h2 class="section-title">📖 Référence légale</h2>
                    <div class="ref-box">
                        {ref_text}
                    </div>
                </div>
                
                <div class="section">
                    <h2 class="section-title">📊 Détail du calcul par période</h2>
                    <table>
                        <thead>
                            <tr>
                                <th>Période</th>
                                <th class="center">Jours</th>
                                <th class="center">Taux BCE</th>
                                <th class="center">Majoration</th>
                                <th class="center">Taux appliqué</th>
                                <th class="right">Intérêts</th>
                            </tr>
                        </thead>
                        <tbody>
                            {rows_html}
                        </tbody>
                    </table>
                </div>
                
                <div class="section">
                    <h2 class="section-title">💰 Récapitulatif</h2>
                    <div class="info-grid">
                        <div class="info-item">
                            <div class="label">Intérêts de retard</div>
                            <div class="value">{format_currency(resultat.interets_totaux)}</div>
                        </div>
                        <div class="info-item">
                            <div class="label">Indemnité forfaitaire (Art. D.441-5)</div>
                            <div class="value">{format_currency(resultat.indemnite_forfaitaire)}</div>
                        </div>
                    </div>
                    <br>
                    <div class="total-box">
                        <div class="label">Total à réclamer</div>
                        <div class="amount">{format_currency(resultat.total_du)}</div>
                    </div>
                </div>
            </div>
            
            <div class="footer">
                Généré par Calculateur Intérêts de Retard — Eurovia / VINCI Construction<br>
                {datetime.now().strftime('%d/%m/%Y à %H:%M')}
            </div>
        </div>
    </body>
    </html>
    """
    
    return html


def generate_csv_export(resultat: ResultatCalcul, libelle: str) -> str:
    """Génère un export CSV."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    writer.writerow(['CALCUL INTÉRÊTS DE RETARD'])
    writer.writerow(['Libellé', libelle])
    writer.writerow(['Montant principal', str(resultat.montant_principal)])
    writer.writerow(['Type client', 'Privé' if resultat.client_type == ClientType.PRIVE else 'Public'])
    writer.writerow(['Date échéance', resultat.date_echeance.strftime('%d/%m/%Y')])
    writer.writerow(['Date paiement', resultat.date_paiement.strftime('%d/%m/%Y')])
    writer.writerow(['Jours retard', resultat.jours_retard])
    writer.writerow([])
    
    writer.writerow(['DÉTAIL PAR PÉRIODE'])
    writer.writerow(['Début', 'Fin', 'Jours', 'Taux BCE', 'Majoration', 'Taux appliqué', 'Intérêts'])
    for seg in resultat.segments:
        writer.writerow([
            seg.debut.strftime('%d/%m/%Y'),
            seg.fin.strftime('%d/%m/%Y'),
            seg.jours,
            f"{seg.taux_bce:.2f}",
            f"+{seg.majoration:.0f}",
            f"{seg.taux_applique:.2f}",
            str(seg.interets)
        ])
    
    writer.writerow([])
    writer.writerow(['RÉCAPITULATIF'])
    writer.writerow(['Intérêts totaux', str(resultat.interets_totaux)])
    writer.writerow(['Indemnité forfaitaire', str(resultat.indemnite_forfaitaire)])
    writer.writerow(['TOTAL DÛ', str(resultat.total_du)])
    
    return output.getvalue()


def generate_excel_export(resultat: ResultatCalcul, libelle: str) -> bytes:
    """Génère un export Excel (.xlsx) stylisé."""
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Calcul Intérêts"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0B3DAA", end_color="0B3DAA", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # En-tête principal
    sheet["A1"] = "CALCULATEUR D'INTÉRÊTS DE RETARD"
    sheet["A1"].font = Font(bold=True, size=14, color="0B3DAA")
    sheet.merge_cells("A1:G1")
    
    # Infos facture
    sheet["A3"] = "Libellé :"
    sheet["B3"] = libelle
    sheet["A4"] = "Montant principal :"
    sheet["B4"] = float(resultat.montant_principal)
    sheet["B4"].number_format = '#,##0.00 €'
    sheet["A5"] = "Type client :"
    sheet["B5"] = "Privé" if resultat.client_type == ClientType.PRIVE else "Public"
    sheet["A6"] = "Date échéance :"
    sheet["B6"] = resultat.date_echeance
    sheet["B6"].number_format = 'dd/mm/yyyy'
    sheet["A7"] = "Date paiement :"
    sheet["B7"] = resultat.date_paiement
    sheet["B7"].number_format = 'dd/mm/yyyy'
    sheet["A8"] = "Jours de retard :"
    sheet["B8"] = resultat.jours_retard

    for cell in sheet["A3:A8"]:
        cell[0].font = Font(bold=True)
    
    # Tableau - En-têtes
    headers = ["Début", "Fin", "Jours", "Taux BCE", "Majoration", "Taux appliqué", "Intérêts"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=10, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Tableau - Données
    row_idx = 11
    for seg in resultat.segments:
        sheet.cell(row=row_idx, column=1, value=seg.debut).number_format = 'dd/mm/yyyy'
        sheet.cell(row=row_idx, column=2, value=seg.fin).number_format = 'dd/mm/yyyy'
        sheet.cell(row=row_idx, column=3, value=seg.jours)
        sheet.cell(row=row_idx, column=4, value=seg.taux_bce / 100).number_format = '0.00%'
        sheet.cell(row=row_idx, column=5, value=seg.majoration)
        sheet.cell(row=row_idx, column=6, value=seg.taux_applique / 100).number_format = '0.00%'
        sheet.cell(row=row_idx, column=7, value=float(seg.interets)).number_format = '#,##0.00 €'
        
        # Bordures
        for col in range(1, 8):
            sheet.cell(row=row_idx, column=col).border = thin_border
        
        row_idx += 1
    
    # Totaux
    row_idx += 2
    sheet.cell(row=row_idx, column=6, value="Intérêts totaux :").font = Font(bold=True)
    sheet.cell(row=row_idx, column=7, value=float(resultat.interets_totaux)).number_format = '#,##0.00 €'
    
    row_idx += 1
    sheet.cell(row=row_idx, column=6, value="Indemnité forfaitaire :").font = Font(bold=True)
    sheet.cell(row=row_idx, column=7, value=float(resultat.indemnite_forfaitaire)).number_format = '#,##0.00 €'
    
    row_idx += 1
    total_cell_label = sheet.cell(row=row_idx, column=6, value="TOTAL À RÉCLAMER :")
    total_cell_label.font = Font(bold=True, size=12, color="0B3DAA")
    
    total_cell_val = sheet.cell(row=row_idx, column=7, value=float(resultat.total_du))
    total_cell_val.font = Font(bold=True, size=12, color="0B3DAA")
    total_cell_val.number_format = '#,##0.00 €'

    # Ajustement colonnes
    for col in range(1, 8):
        sheet.column_dimensions[get_column_letter(col)].width = 15
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20

    workbook.save(output)
    return output.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# INTERFACE UTILISATEUR
# ═══════════════════════════════════════════════════════════════════════════════

def load_logo_base64(path: str) -> str:
    """Charge un logo en base64."""
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except Exception:
        return ""


def render_header():
    """Affiche l'en-tête de l'application."""
    col1, col2 = st.columns([3, 1])
    
    with col1:
        st.markdown("""
        <div class="main-header">
            <h1>💶 Calculateur d'Intérêts de Retard</h1>
            <p>Eurovia / VINCI Construction — Outil de recouvrement BTP</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        # Logo
        logo_path = "assets/logo.png"
        try:
            st.image(logo_path, width=240)  # Increased from 180
        except Exception:
            pass


def render_sidebar():
    """Affiche la sidebar avec informations et aide."""
    with st.sidebar:
        # Logo
        logo_path = "assets/mon_logo.png"
        try:
            st.image(logo_path, width=200)  # Increased from 150
        except Exception:
            pass
        
        st.markdown("---")
        
        st.markdown("### 📚 Modes de calcul")
        
        with st.expander("🏢 Client Privé (L441-10)", expanded=False):
            st.markdown("""
            **Article L.441-10 du Code de commerce**
            
            - **Taux** : BCE (MRO) + 10 points
            - **Actualisation** : Semestrielle
                - S1 : taux au 1er janvier
                - S2 : taux au 1er juillet
            - **Indemnité** : 40 € forfaitaire
            """)
        
        with st.expander("🏛️ Client Public (R2192-31)", expanded=False):
            st.markdown("""
            **Article R.2192-31 du CCP**
            
            - **Taux** : BCE (MRO) + 8 points
            - **Actualisation** : Annuelle (1er janvier)
            - **Indemnité** : 40 € forfaitaire
            """)
        
        with st.expander("📝 Taux Manuel", expanded=False):
            st.markdown("""
            **Clause contractuelle**
            
            - Taux fixe défini par contrat
            - Pas de référence au BCE
            - Indemnité forfaitaire de 40 €
            """)
        
        st.markdown("---")
        
        # Statut des taux BCE
        st.markdown("### 📡 Taux BCE")
        
        if st.button("🔄 Rafraîchir les taux", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
        
        try:
            bce_rates = get_bce_rates()
            if bce_rates:
                last_rate = bce_rates[-1]
                st.success(f"✅ {len(bce_rates)} taux chargés")
                st.info(f"Dernier taux : **{last_rate['rate']:.2f}%**\nau {last_rate['start'].strftime('%d/%m/%Y')}")
        except Exception as e:
            st.error(f"❌ Erreur: {e}")
        
        st.markdown("---")
        
        st.markdown("""
        <div style="text-align: center; color: #6B7280; font-size: 0.75rem;">
            Version 2.0.0<br>
            © 2026 Eurovia / VINCI
        </div>
        """, unsafe_allow_html=True)


def main():
    """Point d'entrée principal de l'application."""
    
    # Initialisation du state
    if 'resultat' not in st.session_state:
        st.session_state.resultat = None
    if 'historique' not in st.session_state:
        st.session_state.historique = []

    # Localisation (Tentative)
    try:
        import locale
        locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')
    except:
        try:
             locale.setlocale(locale.LC_ALL, 'fra') # Windows
        except:
             pass
    
    # Header et sidebar
    render_header()
    render_sidebar()
    
    # Formulaire principal
    st.markdown("### 📝 Saisie de la facture")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">💳 Informations facture</div>', unsafe_allow_html=True)
        
        libelle = st.text_input(
            "Libellé de la facture",
            placeholder="Ex: Facture 2024-001 - Chantier XYZ"
        )
        
        montant = st.number_input(
            "Montant TTC (€)",
            min_value=0.01,
            value=10000.00,
            step=100.00,
            format="%.2f"
        )
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">📅 Dates</div>', unsafe_allow_html=True)
        
        date_echeance = st.date_input(
            "Date d'échéance",
            value=date.today() - timedelta(days=60),
            format="DD/MM/YYYY"
        )
        
        # Calculateur d'échéance (si Public) : Optionnel mais demandé pour les marchés publics
        with st.expander("🛠️ Calculateur d'échéance (Marchés Publics)"):
            mp_date_facture = st.date_input("Date de réception facture", value=date.today(), key="mp_date_fact", format="DD/MM/YYYY")
            mp_delai = st.selectbox("Délai de paiement", options=[30, 50, 60], format_func=lambda x: f"{x} jours")
            if st.button("Appliquer cette échéance", key="btn_apply_date"):
                calculated_due_date = mp_date_facture + timedelta(days=mp_delai)
                st.success(f"Échéance calculée : {calculated_due_date.strftime('%d/%m/%Y')}")
                # Note: On ne peut pas mettre à jour directement le widget date_input ci-dessus sans rerun/session state complexe. 
                # On affiche juste l'info pour que l'utilisateur la recopie ou on utilise session_state si on refond le widget.
                st.info("Veuillez reporter cette date dans le champ 'Date d'échéance' ci-dessus.")

        date_paiement = st.date_input(
            "Date de paiement effectif",
            value=date.today(),
            format="DD/MM/YYYY"
        )
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Type de client et mode
    st.markdown("### ⚙️ Paramètres de calcul")
    
    col3, col4 = st.columns(2)
    
    with col3:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">👥 Type de client</div>', unsafe_allow_html=True)
        
        client_type = st.radio(
            "Sélectionnez le type de client",
            options=["Privé", "Public"],
            horizontal=True,
            label_visibility="collapsed"
        )
        
        if client_type == "Privé":
            st.info("📌 **L.441-10 C.Com** : BCE + 10 pts (semestriel)")
        else:
            st.info("📌 **R.2192-31 CCP** : BCE + 8 pts (annuel)")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col4:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">📊 Mode de taux</div>', unsafe_allow_html=True)
        
        mode_options = ["Légal (automatique)", "Manuel (contractuel)"]
        mode = st.radio(
            "Mode de calcul",
            options=mode_options,
            horizontal=True,
            label_visibility="collapsed"
        )
        
        taux_manuel = None
        if mode == "Manuel (contractuel)":
            taux_manuel = st.number_input(
                "Taux contractuel (%)",
                min_value=0.01,
                max_value=100.0,
                value=12.0,
                step=0.5,
                format="%.2f"
            )
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Bouton de calcul
    st.markdown("<br>", unsafe_allow_html=True)
    
    col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
    
    with col_btn2:
        calculate_clicked = st.button(
            "🧮 CALCULER LES INTÉRÊTS",
            use_container_width=True,
            type="primary"
        )
    
    # Exécution du calcul
    if calculate_clicked:
        # Validation
        if date_paiement <= date_echeance:
            st.warning("⚠️ Le paiement a été effectué dans les délais. Aucun intérêt de retard n'est dû.")
            st.session_state.resultat = None
        else:
            with st.spinner("Calcul en cours..."):
                try:
                    montant_decimal = Decimal(str(montant))
                    
                    if mode == "Manuel (contractuel)":
                        resultat = calculer_interets_manuel(
                            montant_decimal,
                            date_echeance,
                            date_paiement,
                            taux_manuel
                        )
                    elif client_type == "Privé":
                        bce_rates = get_bce_rates()
                        if not bce_rates:
                            st.error("❌ Impossible de charger les taux BCE. Veuillez réessayer.")
                            st.stop()
                        resultat = calculer_interets_prive_l441_10(
                            montant_decimal,
                            date_echeance,
                            date_paiement,
                            bce_rates
                        )
                    else:  # Public
                        bce_rates = get_bce_rates()
                        if not bce_rates:
                            st.error("❌ Impossible de charger les taux BCE. Veuillez réessayer.")
                            st.stop()
                        resultat = calculer_interets_public_r2192_31(
                            montant_decimal,
                            date_echeance,
                            date_paiement,
                            bce_rates
                        )
                    
                    st.session_state.resultat = resultat
                    st.session_state.libelle = libelle or "Facture sans libellé"
                    
                    # Ajouter à l'historique
                    st.session_state.historique.append({
                        'libelle': libelle or "Facture sans libellé",
                        'montant': montant,
                        'total': float(resultat.total_du),
                        'date': datetime.now().strftime('%d/%m/%Y %H:%M')
                    })
                    
                except Exception as e:
                    st.error(f"❌ Erreur lors du calcul : {e}")
    
    # Affichage des résultats
    if st.session_state.resultat:
        resultat = st.session_state.resultat
        libelle = st.session_state.get('libelle', 'Facture')
        
        st.markdown("---")
        st.markdown("### 📊 Résultats du calcul")
        
        # Métriques principales
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        
        with col_m1:
            st.metric(
                label="Montant principal",
                value=format_currency(resultat.montant_principal)
            )
        
        with col_m2:
            st.metric(
                label="Jours de retard",
                value=f"{resultat.jours_retard} jours"
            )
        
        with col_m3:
            st.metric(
                label="Intérêts de retard",
                value=format_currency(resultat.interets_totaux)
            )
        
        with col_m4:
            st.markdown(f"""
            <div class="result-card success">
                <div class="result-label">TOTAL À RÉCLAMER</div>
                <div class="result-value">{format_currency(resultat.total_du)}</div>
            </div>
            """, unsafe_allow_html=True)
        
        # Détail des segments
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("#### 📋 Détail par période")
        
        if resultat.segments:
            df_data = []
            for seg in resultat.segments:
                df_data.append({
                    'Période': f"{seg.debut.strftime('%d/%m/%Y')} → {seg.fin.strftime('%d/%m/%Y')}",
                    'Référence': seg.periode_label,
                    'Jours': seg.jours,
                    'Taux BCE': f"{seg.taux_bce:.2f} %" if seg.taux_bce > 0 else "-",
                    'Majoration': f"+{seg.majoration:.0f} pts" if seg.majoration > 0 else "-",
                    'Taux appliqué': f"{seg.taux_applique:.2f} %",
                    'Intérêts': format_currency(seg.interets)
                })
            
            df = pd.DataFrame(df_data)
            st.dataframe(df, use_container_width=True, hide_index=True)
        
        # Récapitulatif
        col_recap1, col_recap2 = st.columns(2)
        
        with col_recap1:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="card-title">💰 Récapitulatif</div>', unsafe_allow_html=True)
            
            st.markdown(f"""
            | Poste | Montant |
            |-------|---------|
            | Intérêts de retard | {format_currency(resultat.interets_totaux)} |
            | Indemnité forfaitaire (Art. D.441-5) | {format_currency(resultat.indemnite_forfaitaire)} |
            | **TOTAL** | **{format_currency(resultat.total_du)}** |
            """)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col_recap2:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<div class="card-title">📤 Exports</div>', unsafe_allow_html=True)
            
            col_exp1, col_exp2 = st.columns(2)
            
            with col_exp1:
                html_report = generate_html_report(resultat, libelle)
                st.download_button(
                    "📄 Télécharger HTML",
                    data=html_report,
                    file_name=f"interets_retard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html",
                    mime="text/html",
                    use_container_width=True
                )
            
            with col_exp2:
                excel_data = generate_excel_export(resultat, libelle)
                st.download_button(
                    "📊 Télécharger EXCEL",
                    data=excel_data,
                    file_name=f"interets_retard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Référence légale
        st.markdown("<br>", unsafe_allow_html=True)
        with st.expander("📖 Référence légale applicable"):
            if resultat.taux_mode == TauxMode.L441_10:
                st.markdown("""
                **Article L.441-10 du Code de commerce**
                
                Les pénalités de retard sont exigibles sans qu'un rappel soit nécessaire. 
                Sauf disposition contraire qui ne peut toutefois fixer un taux inférieur 
                à trois fois le taux d'intérêt légal, le taux des pénalités de retard 
                est égal au taux d'intérêt appliqué par la Banque centrale européenne à 
                son opération de refinancement la plus récente majoré de **10 points de pourcentage**.
                
                Le taux applicable pendant le premier semestre de l'année concernée est 
                le taux en vigueur au 1er janvier de l'année en question. Pour le second 
                semestre de l'année concernée, il est le taux en vigueur au 1er juillet 
                de l'année en question.
                """)
            elif resultat.taux_mode == TauxMode.R2192_31:
                st.markdown("""
                **Article R.2192-31 du Code de la commande publique**
                
                Le taux des intérêts moratoires est égal au taux d'intérêt de la principale 
                facilité de refinancement appliquée par la Banque centrale européenne à son 
                opération de refinancement principal la plus récente effectuée avant le 
                premier jour de calendrier du semestre de l'année civile au cours duquel 
                les intérêts moratoires ont commencé à courir, majoré de **8 points de pourcentage**.
                """)
            else:
                st.markdown("""
                **Clause contractuelle**
                
                Le taux d'intérêt de retard est celui défini par les parties dans le contrat.
                L'indemnité forfaitaire de 40 € pour frais de recouvrement reste applicable 
                conformément à l'article D.441-5 du Code de commerce.
                """)
    
    # Historique
    if st.session_state.historique:
        st.markdown("---")
        st.markdown("### 📜 Historique des calculs")
        
        df_hist = pd.DataFrame(st.session_state.historique[::-1][:10])  # 10 derniers
        df_hist.columns = ['Libellé', 'Montant (€)', 'Total dû (€)', 'Date calcul']
        st.dataframe(df_hist, use_container_width=True, hide_index=True)
        
        if st.button("🗑️ Effacer l'historique"):
            st.session_state.historique = []
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# POINT D'ENTRÉE
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    main()
